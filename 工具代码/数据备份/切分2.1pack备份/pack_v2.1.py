import os
import re
import time
import paramiko
import threading
import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime
from typing import Optional
import stat
import openpyxl  # 新增依赖
import sys
import tempfile

def clean_ansi(text):
    """清理ANSI转义序列"""
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])|\\033\[[0-9;]*[mK]')
    return ansi_escape.sub('', text)

class ProgressBar(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(fg_color="transparent")
        
        # 创建进度条和文字的容器
        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="x", padx=1, pady=0)
        
        # 进度文本（在进度条上方）
        self.label = ctk.CTkLabel(
            self.container,
            text="准备就绪",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            text_color="#666666",
            bg_color="transparent"
        )
        self.label.pack(pady=(0, 3))  # 减小文字和进度条的间距
        
        # 进度条
        self.progress = ctk.CTkProgressBar(self.container)
        self.progress.pack(fill="x", pady=0)
        self.progress.configure(
            fg_color="#f0f2f5",  # 更柔和的背景色
            progress_color="#4a90e2",  # 更现代的蓝色
            height=18,  # 适中的高度
            border_color="#e0e3e7",  # 更柔和的边框色
            border_width=1,
            corner_radius=4  # 更小的圆角
        )
        self.progress.set(0)
        
        # 详细进度文本（在进度条下方）
        self.detail_label = ctk.CTkLabel(
            self.container,
            text="",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=11),
            text_color="#888888",
            bg_color="transparent"
        )
        self.detail_label.pack(pady=(2, 0))
        
        self._current_value = 0
        self._current_phase = "准备"
        
    def _format_progress(self, value):
        """格式化进度值显示"""
        return f"{value:.1f}%" if value < 100 else "100%"
        
    def _get_status_emoji(self, value):
        """根据进度值返回对应的状态图标"""
        if value == 0:
            return "⚪"  # 未开始
        elif value < 30:
            return "🔄"  # 开始阶段
        elif value < 60:
            return "⏳"  # 处理中
        elif value < 90:
            return "📦"  # 数据处理
        elif value < 100:
            return "📤"  # 传输阶段
        else:
            return "✅"  # 完成
            
    def _get_phase_description(self, value, phase):
        """根据进度值和阶段返回详细描述"""
        if value == 0:
            return "等待开始..."
        elif value == 100:
            return "任务已完成"
        else:
            phase_descriptions = {
                "准备": "正在准备环境...",
                "连接": "正在连接服务器...",
                "检查": "正在检查程序...",
                "配置": "正在配置环境...",
                "切分": "正在切分数据...",
                "复制": "正在复制文件...",
                "清理": "正在清理临时文件...",
                "错误": "处理出现错误"
            }
            return phase_descriptions.get(phase, "正在处理...")
            
    def update_progress(self, value: float, text: Optional[str] = None, phase: Optional[str] = None):
        """更新进度条和文本"""
        # 确保进度值不会倒退（除非是新的操作）
        if value < self._current_value and self._current_value < 1 and not text:
            return
            
        # 确保进度值在0-1之间
        value = max(0, min(1, value))
        self._current_value = value
        
        if phase:
            self._current_phase = phase
            
        # 计算百分比
        percentage = value * 100
        
        # 设置进度条颜色
        if percentage == 100:
            self.progress.configure(progress_color="#4caf50")  # 完成时显示绿色
        elif text and "错误" in text:
            self.progress.configure(progress_color="#f44336")  # 错误时显示红色
        else:
            self.progress.configure(progress_color="#4a90e2")  # 默认蓝色
            
        # 更新进度条
        self.progress.set(value)
        
        # 更新主状态文本
        status_text = text if text else self._get_phase_description(percentage, self._current_phase)
        emoji = self._get_status_emoji(percentage)
        self.label.configure(text=f"{emoji} {status_text}")
        
        # 更新详细进度文本
        if text and "错误" in text:
            self.detail_label.configure(text=f"❌ {text}", text_color="#f44336")
        else:
            detail_text = f"进度：{self._format_progress(percentage)}"
            if phase:
                detail_text += f" | 阶段：{phase}"
            self.detail_label.configure(text=detail_text, text_color="#888888")

class StatusBadge(ctk.CTkLabel):
    def __init__(self, master, **kwargs):
        super().__init__(
            master,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12, weight="bold"),
            corner_radius=18,
            **kwargs
        )
        
    def set_status(self, status: str):
        """设置状态标签的样式"""
        if status == "running":
            self.configure(
                text="运行中...",
                fg_color="#1976d2",  # 深蓝色
                text_color="#ffffff"
            )
        elif status == "success":
            self.configure(
                text="完 成",
                fg_color="#4caf50",  # 绿色
                text_color="#ffffff"
            )
        elif status == "error":
            self.configure(
                text="错 误",
                fg_color="#f44336",  # 红色
                text_color="#ffffff"
            )
        elif status == "idle":
            self.configure(
                text="就 绪",
                fg_color="#FFB6C1",  # 粉色
                text_color="#ffffff"
            )
        elif status == "warning":
            self.configure(
                text="警 告",
                fg_color="#ff9800",  # 橙色
                text_color="#ffffff"
            )
        elif status == "processing":
            self.configure(
                text="处理中",
                fg_color="#2196f3",  # 蓝色
                text_color="#ffffff"
            )
        elif status == "waiting":
            self.configure(
                text="等待中",
                fg_color="#9c27b0",  # 紫色
                text_color="#ffffff"
            )

class HelpDialog(ctk.CTkToplevel):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        
        self.withdraw()
        self.title("帮助")
        self.resizable(False, False)
        self.transient(parent)
        
        # 设置窗口大小和位置
        window_width = 382
        window_height = 320
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 创建内容框架
        self.content_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 创建文本框
        self.help_text = ctk.CTkTextbox(
            self.content_frame,
            font=ctk.CTkFont(family="Consolas", size=12),
            fg_color="#f8f9fa",
            text_color="#333333",
            border_color="#dee2e6",
            border_width=1,
            corner_radius=6
        )
        self.help_text.pack(fill="both", expand=True, pady=(0, 15))
        
        # 帮助文本内容
        help_content = """
=================================================
                CARIZON 数据切分工具
=================================================

【功能简介】
-------------------------------------------------
版本号：V7.1.5
开发人员：CARIZON-Kangming.Fang

【功能说明】
本工具用于将指定数据ID的数据进行切片处理，并支持自定义存储位置、进度监控和远程终止。

【操作流程】
-------------------------------------------------
1. 输入数据ID
   - 在“数据ID”输入框填写需切分的数据ID。
   - 支持导入表格用于快捷识别数据ID，格式为xlsx/txt。
2. 设置切分时间
   - “向前时间”：指定切分点前的时长（秒）。
   - “向后时间”：指定切分点后的时长（秒）。
3. 选择存储位置
   - 可选择“默认地址”或“自定义地址”。
   - 选择自定义时，需输入或浏览选择本地存储路径。
4. 启动任务
   - 点击“开始切分”按钮，任务开始，相关控件自动禁用，防止误操作。
   - 任务运行中，进度和状态实时显示。
5. 终止任务
   - 切分任务运行中可点击“终止程序”按钮，系统弹窗确认，确认后终止任务。
   - 未运行时点击“终止程序”按钮无响应。
6. 任务结束
   - 任务完成或终止后，控件自动恢复，可继续操作。
7. 日志管理
   - 可导出或清空日志，便于追踪操作记录。

【注意事项】
-------------------------------------------------
* 切分任务运行中，存储位置、数据ID、导入按钮等控件均不可更改。
* 终止任务仅终止切分进程，不会清理NAS/OUT目录数据。
* 任务异常或连接失败时，系统自动恢复控件并给出详细提示。
* 请确保网络连接稳定、存储空间充足。
* 数据ID需符合规范格式，否则无法参与切分。
"""
        # 插入文本内容
        self.help_text.insert("1.0", help_content.strip())
        self.help_text.configure(state="disabled")
        
        # 关闭按钮
        close_button = ctk.CTkButton(
            self.content_frame,
            text="关闭",
            width=100,
            height=32,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=13),
            fg_color="#1976d2",
            hover_color="#1565c0",
            corner_radius=6,
            command=self.destroy
        )
        close_button.pack()
        
        self.after(100, self.show_dialog)
        
    def show_dialog(self):
        self.update_idletasks()
        self.deiconify()

class DataSlicerApp:
    def __init__(self):
        # 设置全局主题
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        
        # 创建主窗口
        self.app = ctk.CTk()
        
        # 配置窗口
        self.app.title("CARIZON 数据切分工具")
        self.app.minsize(930, 620)
        self.app.maxsize(930, 620)  # 设置最大尺寸等于最小尺寸，禁止调整大小
        self.app.resizable(False, False)  # 禁止调整大小，包括最大化
        
        # 设置窗口大小和位置
        window_width = 920
        window_height = 600
        screen_width = self.app.winfo_screenwidth()
        screen_height = self.app.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.app.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # SSH配置
        self.ssh_config = {
            "hostname": "10.22.33.86",
            "username": "user",
            "password": "123456",
            "script_path": "/home/user/fkm/pack/new_pack.sh"
        }
        
        # 设置UI
        self.setup_ui()
        
        # 状态变量
        self.is_running = False
        
        # 初始化进度条状态
        self.progress_value = 0
        self.current_phase = ""
        
        # 设置默认存储路径
        self.default_storage_path = r"\\10.21.25.201\Data"
        # 检查Auto_pack.sh路径，若为打包环境则释放到临时目录
        self.auto_pack_sh_path = self._get_auto_pack_sh_path()

    def _get_auto_pack_sh_path(self):
        """获取Auto_pack.sh的实际路径，支持PyInstaller打包环境"""
        import shutil
        script_name = "Auto_pack.sh"
        meipass = getattr(sys, '_MEIPASS', None)
        if meipass:
            # PyInstaller打包环境，资源在_MEIPASS目录
            src = os.path.join(meipass, script_name)
            dst = os.path.join(tempfile.gettempdir(), script_name)
            if not os.path.exists(dst):
                shutil.copy(src, dst)
            return dst
        else:
            # 普通开发环境
            return os.path.join(os.path.dirname(os.path.abspath(__file__)), script_name)

    def setup_ui(self):
        # 主框架
        main_frame = ctk.CTkFrame(self.app, fg_color="#f5f7fa")
        main_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        # 顶部标题栏
        title_bar = ctk.CTkFrame(main_frame, fg_color="#e3eaf2", height=48)
        title_bar.pack(fill="x", pady=(0, 15))
        title_bar.pack_propagate(False)

        # 标题和版本号
        title_frame = ctk.CTkFrame(title_bar, fg_color="transparent")
        title_frame.pack(side="left", padx=20)

        title_label = ctk.CTkLabel(
            title_frame,
            text="CARIZON",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=20, weight="bold"),
            text_color="#FFA500"
        )
        title_label.pack(side="left")

        version_label = ctk.CTkLabel(
            title_frame,
            text="",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            text_color="#666666"
        )
        version_label.pack(side="left", padx=(8, 0), pady=(3, 0))

        # 状态标签
        self.status_badge = StatusBadge(title_frame)
        self.status_badge.pack(side="left", padx=(15, 0))
        self.status_badge.set_status("idle")

        # 帮助按钮
        help_button = ctk.CTkButton(
            title_bar,
            text="帮助",
            width=65,
            height=25,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            fg_color="#1976d2",
            hover_color="#1565c0",
            corner_radius=6,
            command=self.show_help
        )
        help_button.pack(side="right", padx=14, pady=10)

        # 内容区域
        content_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        content_frame.pack(fill="both", expand=True)
        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_columnconfigure(1, weight=2)
        content_frame.grid_rowconfigure(0, weight=1)

        # 左侧参数区
        left_panel = ctk.CTkFrame(content_frame, fg_color="#ffffff", corner_radius=8)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)

        # 右侧状态区
        right_panel = ctk.CTkFrame(content_frame, fg_color="#ffffff", corner_radius=8)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)

        # 左侧内容布局
        self.setup_left_panel(left_panel)
        
        # 右侧内容布局
        self.setup_right_panel(right_panel)

    def setup_left_panel(self, panel):
        """设置左侧面板的内容"""
        # 数据ID输入区域
        id_section = ctk.CTkFrame(panel, fg_color="transparent")
        id_section.pack(fill="x", padx=15, pady=15)

        id_header = ctk.CTkFrame(id_section, fg_color="transparent")
        id_header.pack(fill="x")

        id_label = ctk.CTkLabel(
            id_header,
            text="切分数据ID",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=14, weight="bold"),
            text_color="#1976d2"
        )
        id_label.pack(side="left")

        id_hint = ctk.CTkLabel(
            id_header,
            text="",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            text_color="#666666"
        )
        id_hint.pack(side="left", padx=(8, 0), pady=(1, 0))

        # 新增：导入表格按钮
        self.import_btn = ctk.CTkButton(
            id_header,
            text="导入表格",
            width=80,
            height=26,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            fg_color="#4caf50",
            hover_color="#388e3c",
            corner_radius=6,
            command=self.import_xlsx_ids
        )
        self.import_btn.pack(side="right", padx=(10, 0))

        self.data_id_entry = ctk.CTkTextbox(
            id_section,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            fg_color="#f5f7fa",
            border_color="#b0c4de",
            border_width=1,
            corner_radius=6,
            height=190,
            width=260  # 调整宽度为200
        )
        self.data_id_entry.pack(fill="x", pady=(8, 0))

        # 时间参数区域
        time_section = ctk.CTkFrame(panel, fg_color="transparent")
        time_section.pack(fill="x", padx=15, pady=(0, 15))

        time_label = ctk.CTkLabel(
            time_section,
            text="时间参数",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=14, weight="bold"),
            text_color="#1976d2"
        )
        time_label.pack(anchor="w")

        time_params_frame = ctk.CTkFrame(time_section, fg_color="transparent")
        time_params_frame.pack(fill="x", pady=(8, 0))

        # 向前时间
        forward_frame = ctk.CTkFrame(time_params_frame, fg_color="transparent")
        forward_frame.pack(fill="x", pady=(0, 8))

        forward_label = ctk.CTkLabel(
            forward_frame,
            text="向前时间(s)：",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=13),
            text_color="#333333"
        )
        forward_label.pack(side="left")

        self.forward_entry = ctk.CTkEntry(
            forward_frame,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=11),
            fg_color="#f5f7fa",
            border_color="#b0c4de",
            border_width=1,
            corner_radius=6,
            width=80,
            height=28
        )
        self.forward_entry.pack(side="right")
        self.forward_entry.insert(0, "5")

        # 向后时间
        backward_frame = ctk.CTkFrame(time_params_frame, fg_color="transparent")
        backward_frame.pack(fill="x")

        backward_label = ctk.CTkLabel(
            backward_frame,
            text="向后时间(s)：",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=13),
            text_color="#333333"
        )
        backward_label.pack(side="left")

        self.backward_entry = ctk.CTkEntry(
            backward_frame,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=11),
            fg_color="#f5f7fa",
            border_color="#b0c4de",
            border_width=1,
            corner_radius=6,
            width=80,
            height=28
        )
        self.backward_entry.pack(side="right")
        self.backward_entry.insert(0, "5")

        # 存储位置区域
        storage_section = ctk.CTkFrame(panel, fg_color="transparent")
        storage_section.pack(fill="x", padx=15, pady=(0, 15))

        storage_label = ctk.CTkLabel(
            storage_section,
            text="存储位置",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=14, weight="bold"),
            text_color="#1976d2"
        )
        storage_label.pack(anchor="w")

        self.storage_var = ctk.StringVar(value="1")
        radio_style = {
            "font": ctk.CTkFont(family="Microsoft YaHei UI", size=13),
            "fg_color": "#1976d2",
            "border_color": "#1976d2",
            "text_color": "#333333",
            "hover_color": "#1565c0"
        }

        radio_frame = ctk.CTkFrame(storage_section, fg_color="transparent")
        radio_frame.pack(fill="x", pady=(8, 0))

        self.radio_default = ctk.CTkRadioButton(
            radio_frame,
            text="默认地址",
            variable=self.storage_var,
            value="1",
            command=self.toggle_custom_path,
            **radio_style
        )
        self.radio_default.pack(side="left", padx=(0, 25))

        self.radio_custom = ctk.CTkRadioButton(
            radio_frame,
            text="自定义地址",
            variable=self.storage_var,
            value="2",
            command=self.toggle_custom_path,
            **radio_style
        )
        self.radio_custom.pack(side="left")

        # 自定义路径框架 - 始终显示
        self.custom_path_frame = ctk.CTkFrame(storage_section, fg_color="transparent")
        self.custom_path_frame.pack(fill="x", pady=(10, 0))

        path_input_frame = ctk.CTkFrame(self.custom_path_frame, fg_color="transparent")
        path_input_frame.pack(fill="x")

        self.custom_path_entry = ctk.CTkEntry(
            path_input_frame,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=11),
            placeholder_text="请输入或选择存储路径",
            fg_color="#f5f7fa",
            border_color="#b0c4de",
            border_width=1,
            corner_radius=6,
            height=28,
            text_color="#333333",
            state="disabled"  # 默认禁用状态
        )
        self.custom_path_entry.pack(side="left", fill="x", expand=True)

        self.browse_button = ctk.CTkButton(
            path_input_frame,
            text="浏览",
            width=70,
            height=28,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=13),
            fg_color="#1976d2",
            hover_color="#1565c0",
            corner_radius=6,
            command=self.browse_directory,
            state="disabled"  # 默认禁用状态
        )
        self.browse_button.pack(side="right", padx=(8, 0))

        # 执行按钮
        self.run_button = ctk.CTkButton(
            panel,
            text="开始执行",
            command=self.run_script,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=14, weight="bold"),
            height=36,
            fg_color="#1976d2",
            hover_color="#1565c0",
            corner_radius=6
        )
        self.run_button.pack(fill="x", padx=15, pady=(0, 15))

    def setup_right_panel(self, panel):
        """设置右侧面板的内容"""
        # 标题栏和清空按钮容器
        header_frame = ctk.CTkFrame(panel, fg_color="transparent")
        header_frame.pack(fill="x", padx=15, pady=12)
        
        # 状态标题和图标
        status_header = ctk.CTkFrame(header_frame, fg_color="transparent")
        status_header.pack(side="left")
        
        status_label = ctk.CTkLabel(
            status_header,
            text="执行状态",
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=14, weight="bold"),
            text_color="#1976d2"
        )
        status_label.pack(side="left")
        
        # 清空和导出按钮框架
        buttons_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        buttons_frame.pack(side="right")
        
        # 导出日志按钮
        export_button = ctk.CTkButton(
            buttons_frame,
            text="导出日志",
            width=65,
            height=25,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            fg_color="#1976d2",
            hover_color="#1565c0",
            corner_radius=6,
            command=self.export_log
        )
        # 新增：终止程序按钮
        self.terminate_button = ctk.CTkButton(
            buttons_frame,
            text="终止程序",
            width=65,
            height=25,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            fg_color="#e91e63",
            hover_color="#c2185b",
            corner_radius=6,
            command=self.on_terminate_button_click  # 修改为新的回调
        )
        self.terminate_button.pack(side="left", padx=(0, 8))
        export_button.pack(side="left", padx=(0, 8))
        
        # 清空按钮
        clear_button = ctk.CTkButton(
            buttons_frame,
            text="清空日志",
            width=65,
            height=25,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=12),
            fg_color="#e91e63",
            hover_color="#c2185b",
            corner_radius=6,
            command=self.clear_status
        )
        clear_button.pack(side="left")
        
        # 状态文本区域
        status_frame = ctk.CTkFrame(panel, fg_color="transparent")
        status_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # 状态文本框
        self.status_text = ctk.CTkTextbox(
            status_frame,
            font=ctk.CTkFont(family="Microsoft YaHei UI", size=11),
            fg_color="#f5f7fa",
            border_color="#b0c4de",
            border_width=1,
            corner_radius=6,
            text_color="#333333",
            state="disabled"
        )
        self.status_text.pack(fill="both", expand=True)

    def show_help(self):
        """显示帮助对话框"""
        self.help_dialog = HelpDialog(self.app)
        self.help_dialog.focus()
        # 设置为模态对话框
        self.help_dialog.grab_set()

    def import_xlsx_ids(self):
        """导入xlsx或txt文件，提取去重后的数据ID，填入数据ID输入框"""
        file_path = filedialog.askopenfilename(
            title="选择Excel或TXT文件",
            filetypes=[("Excel files", "*.xlsx"), ("Text files", "*.txt"), ("All files", "*.*")]
        )
        if not file_path:
            return
        import re, os
        id_set = set()
        pattern = re.compile(r"[A-Za-z0-9]{6}_[0-9]{8}-[0-9]{6}-[0-9]{3}")
        try:
            if file_path.lower().endswith('.xlsx'):
                file_name = os.path.basename(file_path)
                self.update_status(f"开始导入Excel表格: {file_name}", "info")
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    max_row = ws.max_row
                    max_col = ws.max_column
                    sheet_id_set = set()
                    for row_idx in range(1, max_row + 1):
                        for col_idx in range(1, max_col + 1):
                            try:
                                cell = ws.cell(row=row_idx, column=col_idx).value
                                if cell is None:
                                    continue
                                cell_str = str(cell).strip().replace('\u200b', '').replace('\ufeff', '').replace('\xa0', '')
                                for match in pattern.findall(cell_str):
                                    id_set.add(match)
                                    sheet_id_set.add(match)
                            except Exception as ce:
                                self.update_status(f"单元格处理异常: {ce}", "warning")
                    self.update_status(f"正在遍历Excel表格: 共识别：{len(sheet_id_set)}个数据ID", "info")
            elif file_path.lower().endswith('.txt'):
                file_name = os.path.basename(file_path)
                self.update_status(f"开始导入txt文件: {file_name}", "info")
                txt_id_set = set()
                with open(file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        for match in pattern.findall(line):
                            id_set.add(match)
                            txt_id_set.add(match)
                self.update_status(f"正在遍历txt文件:共识别：{len(txt_id_set)}个数据ID", "info")
            else:
                messagebox.showwarning("提示", "仅支持xlsx或txt文件！")
                return
            if not id_set:
                self.update_status("未在文件中找到任何符合格式的数据ID！", "warning")
                messagebox.showwarning("提示", "未在文件中找到任何符合格式的数据ID！")
                return
            id_str = " ".join(sorted(id_set))
            self.data_id_entry.delete("1.0", "end")
            self.data_id_entry.insert("1.0", id_str)
            messagebox.showinfo("成功", f"已导入{len(id_set)}个数据ID！")
        except Exception as e:
            self.update_status(f"导入文件失败：{str(e)}", "error")
            messagebox.showerror("错误", f"导入文件失败：{str(e)}")

    def browse_directory(self):
        """打开目录选择对话框"""
        directory = filedialog.askdirectory()
        if directory:
            self.custom_path_entry.delete(0, "end")
            self.custom_path_entry.insert(0, directory)
            # 自动选择自定义地址选项
            self.storage_var.set("2")
            self.toggle_custom_path()

    def toggle_custom_path(self):
        """切换自定义路径输入框的状态"""
        # 保存当前输入框的内容
        if not hasattr(self, '_saved_path'):
            self._saved_path = ""
            
        if self.storage_var.get() == "2":
            # 切换到自定义地址
            self.custom_path_entry.configure(state="normal")
            self.browse_button.configure(state="normal")
            # 恢复之前保存的路径
            if self._saved_path:
                self.custom_path_entry.delete(0, "end")
                self.custom_path_entry.insert(0, self._saved_path)
        else:
            # 切换到默认地址
            # 保存当前路径
            self._saved_path = self.custom_path_entry.get()
            # 清空并禁用输入框
            self.custom_path_entry.delete(0, "end")
            self.custom_path_entry.configure(state="disabled")
            self.browse_button.configure(state="disabled")

    def validate_inputs(self):
        # 验证数据ID
        data_ids_raw = self.data_id_entry.get("1.0", "end-1c")
        import re
        id_list = re.split(r'[\s,;]+', data_ids_raw)
        id_list = [i.strip() for i in id_list if i.strip()]
        pattern = re.compile(r"^[A-Za-z0-9]{6}_[0-9]{8}-[0-9]{6}-[0-9]{3}$")
        valid_ids = [i for i in id_list if pattern.match(i)]
        invalid_ids = [i for i in id_list if not pattern.match(i)]
        if not valid_ids:
            self.update_status("❌ 输入格式错误: 数据ID均不符合要求，程序终止。", "error")
            return False
        if invalid_ids:
            self.update_status(f"⚠️ 以下数据ID格式不正确，请稍后检查: {' '.join(invalid_ids)}", "warning")
        # 只用合规ID
        id_str = " ".join(sorted(set(valid_ids)))
        self.data_id_entry.delete("1.0", "end")
        self.data_id_entry.insert("1.0", id_str)
        # 验证时间参数
        forward_time = self.forward_entry.get().strip()
        backward_time = self.backward_entry.get().strip()
        if not forward_time or not backward_time:
            messagebox.showerror("错误", "请输入向前和向后的时间参数")
            return False
        if not forward_time.isdigit() or not backward_time.isdigit():
            messagebox.showerror("错误", "时间参数必须是正整数")
            return False
        # 验证自定义路径
        if self.storage_var.get() == "2" and not self.custom_path_entry.get().strip():
            messagebox.showerror("错误", "请输入自定义存储路径")
            return False
        # 保存本次不合规ID，供后续汇总显示
        self._invalid_ids = invalid_ids
        return True

    def show_invalid_ids_summary(self):
        if hasattr(self, '_invalid_ids') and self._invalid_ids:
            self.update_status(f"⚠️ 本次未参与切分的无效数据ID: {' '.join(self._invalid_ids)}", "warning", show_timestamp=False)

    def convert_path_format(self, linux_path):
        """将Linux格式的路径转换为Windows网络共享格式
        例如：/home/user/10.21.25.201/E81677/... -> \\10.21.25.201\Data\E81677\...
        """
        try:
            if "/home/user/10.21.25.201/" in linux_path:
                # 移除/home/user/前缀，替换为网络路径
                relative_path = linux_path.split("/home/user/10.21.25.201/")[1]
                windows_path = f"{self.default_storage_path}\\{relative_path}"
                return windows_path.replace('/', '\\')
            return linux_path
        except Exception:
            return linux_path  # 转换失败时返回原路径

    def update_status(self, message, level="info", show_timestamp=True):
        # 确保在主线程中更新UI
        def _update():
            # 清理ANSI转义序列
            cleaned_message = clean_ansi(message)
            
            # 添加时间戳
            if show_timestamp:
                timestamp = time.strftime("%H:%M:%S", time.localtime())
                formatted_message = f"[{timestamp}] {cleaned_message}\n"
            else:
                formatted_message = f"{cleaned_message}\n"
            
            # 临时启用编辑
            self.status_text.configure(state="normal")
            self.status_text.insert("end", formatted_message)
            
            # 设置消息颜色
            last_line_start = self.status_text.index("end-2c linestart")
            last_line_end = self.status_text.index("end-1c")
            
            # 根据消息级别设置颜色
            if level == "error":
                self.status_text.tag_add("error", last_line_start, last_line_end)
                self.status_text.tag_config("error", foreground="#f44336")  # 红色
                self.status_badge.set_status("error")
            elif level == "success":
                self.status_text.tag_add("success", last_line_start, last_line_end)
                self.status_text.tag_config("success", foreground="#4caf50")  # 绿色
            elif level == "warning":
                self.status_text.tag_add("warning", last_line_start, last_line_end)
                self.status_text.tag_config("warning", foreground="#ff9800")  # 橙色
            elif level == "processing":
                self.status_text.tag_add("processing", last_line_start, last_line_end)
                self.status_text.tag_config("processing", foreground="#2196f3")  # 蓝色
            elif level == "waiting":
                self.status_text.tag_add("waiting", last_line_start, last_line_end)
                self.status_text.tag_config("waiting", foreground="#9c27b0")  # 紫色
            elif level == "highlight":
                self.status_text.tag_add("highlight", last_line_start, last_line_end)
                self.status_text.tag_config("highlight", foreground="#009688")  # 青色
            elif level == "emphasis":
                self.status_text.tag_add("emphasis", last_line_start, last_line_end)
                self.status_text.tag_config("emphasis", foreground="#3f51b5")  # 靛蓝色
            else:
                # 正常输出使用黑色
                self.status_text.tag_add("normal", last_line_start, last_line_end)
                self.status_text.tag_config("normal", foreground="#000000")  # 黑色
            
            self.status_text.see("end")
            # 恢复只读状态
            self.status_text.configure(state="disabled")
        
        # 如果是主线程直接执行，否则使用after
        try:
            if threading.current_thread() == threading.main_thread():
                _update()
            else:
                self.app.after_idle(_update)
        except Exception as e:
            print(f"更新状态时出错: {e}")

    def update_progress(self, value, text=None, phase=None):
        """线程安全的进度更新 - 已禁用"""
        pass

    def run_script(self):
        if not self.validate_inputs():
            return
        if self.is_running:
            messagebox.showwarning("警告", "任务正在运行中，请等待完成")
            return
        self.is_running = True
        self._terminate_flag = False
        self.run_button.configure(state="disabled", text="正在执行...")
        self.status_badge.set_status("running")
        self.update_status("开始执行数据切分任务...", "info")
        # 置灰控件（新增：存储位置、数据ID栏、导入按钮全部禁用）
        self.forward_entry.configure(state="disabled")
        self.backward_entry.configure(state="disabled")
        self.import_btn.configure(state="disabled")
        self.data_id_entry.configure(state="disabled")
        if hasattr(self, '_storage_var_trace_id') and self._storage_var_trace_id:
            try:
                self.storage_var.trace_vdelete('w', self._storage_var_trace_id)
            except Exception:
                pass
        self.radio_default.configure(state="disabled")
        self.radio_custom.configure(state="disabled")
        self.browse_button.configure(state="disabled")
        self.custom_path_entry.configure(state="disabled")
        # 在新线程中执行SSH操作
        self._worker_thread = threading.Thread(target=self.execute_remote_script, daemon=True)
        self._worker_thread.start()
        
    def execute_remote_script(self):
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            try:
                ssh.connect(
                    self.ssh_config["hostname"],
                    username=self.ssh_config["username"],
                    password=self.ssh_config["password"]
                    # ,port=self.ssh_config.get("port", 22)
                )
            except Exception as e:
                self.update_status(f"❌ SSH连接失败: {e}", "error")
                self.is_running = False
                self.restore_controls()
                return
            
            # 首先检查bolepack的位置
            self.update_progress(0.1, phase="检查")
            self.update_status("正在检查bolepack程序...", "processing")
            check_commands = [
                "find /home/user/fkm -name bolepack -type f 2>/dev/null",
                "find /usr/local -name bolepack -type f 2>/dev/null",
                "find ~/.local -name bolepack -type f 2>/dev/null"
            ]
            
            bolepack_path = None
            for cmd in check_commands:
                stdin, stdout, stderr = ssh.exec_command(cmd)
                result = stdout.read().decode().strip()
                if result:
                    bolepack_path = result.split('\n')[0]
                    self.update_status(f"✅ 找到bolepack路径: {bolepack_path}", "success")
                    break
            
            if not bolepack_path:
                raise Exception("无法找到bolepack程序，请确保程序已正确安装")

            # 检查并设置库文件路径
            self.update_progress(0.15, phase="配置")
            self.update_status("正在配置环境变量...", "processing")
            bolepack_dir = os.path.dirname(bolepack_path)
            lib_paths = [
                f"{bolepack_dir}",
                f"{bolepack_dir}/lib",
                f"{bolepack_dir}/../lib",
                "/home/user/.local/bolepack/lib",
                "/home/user/.local/lib",
                "/home/user/fkm/pack/lib",
                "/usr/local/lib",
                "/usr/lib",
                "/usr/lib/x86_64-linux-gnu",
                "/lib",
                "/lib64"
            ]
            
            # 设置环境变量
            env_setup = (
                f"export PATH={bolepack_dir}:$PATH && "
                f"export LANG=en_US.UTF-8 && "
                f"export LC_ALL=en_US.UTF-8 && "
                f"export LD_LIBRARY_PATH={':'.join(lib_paths)}:$LD_LIBRARY_PATH"
            )
            
            # 检查远程目录是否存在
            self.update_progress(0.2, phase="检查")
            self.update_status("正在检查远程目录...", "processing")
            stdin, stdout, stderr = ssh.exec_command("test -d /home/user/fkm/pack && echo 'exists'")
            if stdout.read().decode().strip() != 'exists':
                self.update_status("⚠️ 远程目录不存在，正在创建...", "warning")
                ssh.exec_command("mkdir -p /home/user/fkm/pack")
                self.update_status("✅ 远程目录创建成功: /home/user/fkm/pack", "success")
            else:
                self.update_status("✅ 远程目录已存在: /home/user/fkm/pack", "success")
            
            # 检查脚本是否存在
            self.update_progress(0.25, phase="检查")
            self.update_status("正在检查切分脚本...", "processing")
            stdin, stdout, stderr = ssh.exec_command("test -f /home/user/fkm/pack/Auto_pack.sh && echo 'exists'")
            if stdout.read().decode().strip() != 'exists':
                self.update_status("⚠️ 正在部署切分程序...", "warning")
                
                # 创建SFTP客户端
                sftp = ssh.open_sftp()
                try:
                    # 上传脚本（使用self.auto_pack_sh_path）
                    local_script_path = self.auto_pack_sh_path
                    remote_script_path = "/home/user/fkm/pack/Auto_pack.sh"
                    sftp.put(local_script_path, remote_script_path)
                    
                    # 设置执行权限
                    ssh.exec_command(f"chmod +x {remote_script_path}")
                    self.update_status("✅ 切分程序部署成功", "success")
                    self.update_status("✅ 已设置脚本执行权限", "success")
                finally:
                    sftp.close()
            else:
                self.update_status("✅ 切分脚本已存在: Auto_pack.sh", "success")
                # 确保脚本有执行权限
                ssh.exec_command(f"chmod +x /home/user/fkm/pack/Auto_pack.sh")
                self.update_status("✅ 已确认脚本执行权限", "success")
            
            # 准备命令
            self.update_progress(0.3, phase="准备")
            self.update_status("🔄 正在准备执行参数...", "processing")
            data_ids = self.data_id_entry.get("1.0", "end-1c").strip()
            storage_option = self.storage_var.get()
            custom_path = self.custom_path_entry.get() if storage_option == "2" else ""
            display_path = custom_path  # 保存用于显示的本地路径
            
            # 如果是自定义路径，在远程创建一个临时目录
            if storage_option == "2":
                remote_temp_dir = f"/tmp/data_slicer_{int(time.time())}"
                self.update_status(f"📁 正在创建远程临时目录: {remote_temp_dir}", "processing")
                
                # 确保本地目录存在
                display_path = os.path.normpath(self.custom_path_entry.get().strip())
                try:
                    os.makedirs(display_path, exist_ok=True)
                    self.update_status(f"✅ 本地目录创建成功: {display_path}", "success")
                except Exception as e:
                    raise Exception(f"创建本地目录失败: {str(e)}")
                
                # 尝试创建远程临时目录
                retry_count = 3
                for i in range(retry_count):
                    try:
                        stdin, stdout, stderr = ssh.exec_command(f"mkdir -p {remote_temp_dir} && test -d {remote_temp_dir} && echo 'success'")
                        result = stdout.read().decode().strip()
                        if result == 'success':
                            self.update_status(f"✅ 远程临时目录创建成功: {remote_temp_dir}", "success")
                            custom_path = remote_temp_dir  # 更新为远程路径
                            break
                        else:
                            if i < retry_count - 1:
                                self.update_status(f"⚠️ 重试创建远程临时目录 ({i+1}/{retry_count})...", "warning")
                                time.sleep(1)
                            else:
                                raise Exception("无法创建远程临时目录")
                    except Exception as e:
                        if i < retry_count - 1:
                            self.update_status(f"⚠️ 重试创建远程临时目录 ({i+1}/{retry_count})...", "warning")
                            time.sleep(1)
                        else:
                            raise Exception(f"创建远程临时目录失败: {str(e)}")

            
            # 构建完整的输入字符串，每个输入后面都加换行
            forward_time = self.forward_entry.get().strip()
            backward_time = self.backward_entry.get().strip()
            
            input_data = (
                f"{data_ids}\n"  # 数据ID
                f"{forward_time}\n"  # 向前时间
                f"{backward_time}\n"  # 向后时间
                f"{storage_option}\n"  # 存储选项
                f"{custom_path}"  # 自定义路径（如果有）
            )
            
            # 构建执行命令
            self.update_progress(0.5, phase="切分")
            self.update_status("🚀 开始执行数据切分...", "highlight")
            self.update_status(f"📊 切分参数 - 向前: {forward_time}秒, 向后: {backward_time}秒", "emphasis")
            if storage_option == "2":
                self.update_status(f"📂 存储位置: {display_path}", "emphasis")  # 使用本地路径显示
            else:
                self.update_status("📂 存储位置: 默认路径", "emphasis")
            
            # 使用 printf 确保输入正确传递，并设置环境变量
            command = f'{env_setup} && printf "%s" "{input_data}" | bash /home/user/fkm/pack/Auto_pack.sh 2>&1'
            
            # 执行命令
            stdin, stdout, stderr = ssh.exec_command(command)
            
            # 实时显示输出，过滤掉locale相关信息
            progress_value = 0.1  # 初始进度值
            current_data_id = None  # 当前正在处理的数据ID
            last_notified_data_id = None  # 上一次已通知的数据ID
            data_id_pattern = re.compile(r'([A-Za-z0-9]{6}_[0-9]{8}-[0-9]{6}-[0-9]{3})')
            slice_results = {}  # 使用字典存储切分结果，以数据ID为键
            
            while True:
                # 检查终止标志
                if getattr(self, '_terminate_flag', False):
                    self.update_status("⚠️ 用户主动终止任务，切分进程已终止，任务中断。", "warning")
                    return  # 立即退出，不再处理后续输出和统计

                line = stdout.readline()
                if not line:
                    break
                
                cleaned_line = line.strip()
                if not cleaned_line:
                    continue
                # 终止后不再输出任何业务提示
                if getattr(self, '_terminate_flag', False):
                    return
                # 新增：捕获远程nas、out目录路径
                nas_dir_match = re.search(r"NAS_DIR:\s*(.+)", cleaned_line)
                if nas_dir_match:
                    self._current_nas_dir = nas_dir_match.group(1).strip()
                out_dir_match = re.search(r"OUT_DIR:\s*(.+)", cleaned_line)
                if out_dir_match:
                    self._current_out_dir = out_dir_match.group(1).strip()

                # 过滤掉locale相关的警告信息和sudo密码提示
                if ("warning" in cleaned_line.lower() and "locale" in cleaned_line.lower()) or \
                   "[sudo] password for user:" in cleaned_line:
                    continue
                
                # 如果输出包含远程临时目录路径，替换为本地路径显示
                if storage_option == "2" and remote_temp_dir in cleaned_line:
                    cleaned_line = cleaned_line.replace(remote_temp_dir, display_path)
                
                # 检查是否有新的数据ID
                match = data_id_pattern.search(cleaned_line)
                if match:
                    current_data_id = match.group(1)
                    if current_data_id not in slice_results:
                        slice_results[current_data_id] = {
                            'data_id': current_data_id,
                            'status': '处理中',
                            'path': None
                        }
                    
                    if current_data_id != last_notified_data_id:
                        self.update_status(f"🔄 正在处理数据ID: {current_data_id}", "highlight")
                        last_notified_data_id = current_data_id
                
                # 更新进度和状态
                if "数据切分已完成" in cleaned_line and current_data_id:
                    progress_value = min(0.9, progress_value + 0.1)
                    self.update_progress(progress_value, phase="切分")
                    slice_results[current_data_id]['status'] = '完成'
                    self.update_status(f"✅ {cleaned_line}", "success")
                elif "错误" in cleaned_line or "失败" in cleaned_line:
                    if current_data_id and current_data_id in slice_results:
                        slice_results[current_data_id]['status'] = '失败'
                    self.update_status(f"❌ {cleaned_line}", "error")
                elif "警告" in cleaned_line or "warning" in cleaned_line.lower():
                    self.update_status(f"⚠️ {cleaned_line}", "warning")
                else:
                    # 检查是否包含切分路径信息
                    if "切分数据地址" in cleaned_line:
                        path_match = re.search(r"切分数据地址:\s*(.+)$", cleaned_line)
                        if path_match:
                            original_path = path_match.group(1).strip()
                            converted_path = self.convert_path_format(original_path)
                            # 更新显示的路径
                            cleaned_line = f"切分数据地址: {converted_path}"
                            if current_data_id and current_data_id in slice_results:
                                slice_results[current_data_id]['path'] = converted_path
                            self.update_status(cleaned_line, "info")
                            continue
                    
                    # 根据不同类型的消息添加不同的图标和颜色
                    if "创建" in cleaned_line:
                        self.update_status(f"📁 {cleaned_line}", "processing")
                    elif "复制" in cleaned_line:
                        self.update_status(f"📋 {cleaned_line}", "processing")
                    elif "清理" in cleaned_line:
                        self.update_status(f"🧹 {cleaned_line}", "processing")
                    elif "检查" in cleaned_line:
                        self.update_status(f"🔍 {cleaned_line}", "processing")
                    elif "完成" in cleaned_line:
                        self.update_status(f"✅ {cleaned_line}", "success")
                    else:
                        self.update_status(f"{cleaned_line}", "info")
            
            # 显示所有切分结果的汇总信息
            if slice_results:
                # 如果是自定义路径，处理文件传输
                if storage_option == "2":
                    self.update_status("\n开始传输数据到本地...", "processing")
                    sftp = ssh.open_sftp()
                    try:
                        # 获取远程目录中的所有文件
                        remote_files = []
                        for root, _, files in self._sftp_walk(sftp, remote_temp_dir):
                            for file in files:
                                remote_path = f"{root}/{file}"
                                # 计算相对路径，以保持目录结构
                                rel_path = os.path.relpath(remote_path, remote_temp_dir)
                                local_path = os.path.join(display_path, rel_path)
                                remote_files.append((remote_path, local_path))

                        # 创建本地目录结构
                        for _, local_path in remote_files:
                            os.makedirs(os.path.dirname(local_path), exist_ok=True)

                        # 传输文件
                        total_files = len(remote_files)
                        for i, (remote_path, local_path) in enumerate(remote_files, 1):
                            try:
                                self.update_status(f"正在传输文件 ({i}/{total_files}): {os.path.basename(local_path)}", "processing")
                                sftp.get(remote_path, local_path)
                            except Exception as e:
                                self.update_status(f"传输文件失败 {remote_path}: {str(e)}", "error")

                        self.update_status("✅ 数据传输完成！", "success")
                        # 更新所有结果的状态为完成
                        for result in slice_results.values():
                            if result['status'] == '处理中':
                                result['status'] = '完成'

                    except Exception as e:
                        self.update_status(f"❌ 数据传输失败：{str(e)}", "error")
                        # 更新所有结果的状态为失败
                        for result in slice_results.values():
                            if result['status'] == '处理中':
                                result['status'] = '失败'
                        raise
                    finally:
                        sftp.close()

                    # 清理远程临时目录
                    try:
                        ssh.exec_command(f"rm -rf {remote_temp_dir}")
                        self.update_status("✅ 已清理远程临时目录", "success")
                    except Exception as e:
                        self.update_status(f"清理远程临时目录失败: {str(e)}", "warning")

                # 显示结果汇总
                self.update_status("\n📋 数据切分结果汇总:", "emphasis", show_timestamp=False)
                result_items = list(slice_results.items())
                for i, (data_id, result) in enumerate(result_items):
                    status_icon = "✅" if result['status'] == '完成' else "❌"
                    status_level = "success" if result['status'] == '完成' else "error"
                    self.update_status(f"{status_icon} 数据ID: {result['data_id']}", status_level, show_timestamp=False)
                    if result['path']:
                        # 确保显示Windows格式的路径
                        display_path = self.convert_path_format(result['path'])
                        self.update_status(f"   📂 切分数据地址: {display_path}", "info", show_timestamp=False)
                    self.update_status(f"   状态: {result['status']}", status_level, show_timestamp=False)
                    # 只在不是最后一个结果时添加空行
                    if i < len(result_items) - 1:
                        self.update_status("", "info", show_timestamp=False)
            
            # 完成所有处理
            self.update_progress(1.0, phase="完成")
            total_data = len(slice_results)
            completed_data = sum(1 for result in slice_results.values() if result['status'] == '完成')
            
            if completed_data == total_data:
                self.update_status(f"✅ 共{total_data}个数据，切分任务已完成", "success", show_timestamp=False)
            else:
                self.update_status(f"⚠️ 共{total_data}个数据，{completed_data}个完成，{total_data - completed_data}个未完成，请检查详细信息", "warning", show_timestamp=False)
            
            ssh.close()
            
            # 线程安全地更新状态标签
            def _update_status():
                try:
                    self.status_badge.set_status("success")
                    # 2秒后将状态更改为就绪
                    self.app.after(2000, lambda: self.status_badge.set_status("idle"))
                except Exception as e:
                    print(f"更新状态时出错: {e}")
            
            # 在主线程中更新状态
            self.app.after_idle(_update_status)
            
        except Exception as e:
            self.update_status(f"❌ 执行出错: {str(e)}", "error")
            # 发生错误时更新状态为错误
            def _update_error_status():
                try:
                    self.status_badge.set_status("error")
                    # 2秒后将状态更改为就绪
                    self.app.after(2000, lambda: self.status_badge.set_status("idle"))
                except:
                    pass
            self.app.after_idle(_update_error_status)
        finally:
            self.is_running = False
            self.restore_controls()
            
    def restore_controls(self):
        self.forward_entry.configure(state="normal")
        self.backward_entry.configure(state="normal")
        self.import_btn.configure(state="normal")
        self.data_id_entry.configure(state="normal")
        self.radio_default.configure(state="normal")
        self.radio_custom.configure(state="normal")
        self._storage_var_trace_id = self.storage_var.trace('w', lambda *args: self.toggle_custom_path())
        if self.storage_var.get() == "2":
            self.custom_path_entry.configure(state="normal")
            self.browse_button.configure(state="normal")
        else:
            self.custom_path_entry.configure(state="disabled")
            self.browse_button.configure(state="disabled")
        self.run_button.configure(state="normal", text="开始执行")

    def terminate_task(self):
        self._terminate_flag = True
        self.update_status("⚠️ 用户主动终止任务，正在强制终止切分进程...", "warning")
        import time
        if hasattr(self, '_worker_thread') and self._worker_thread.is_alive():
            for _ in range(30):
                if not self._worker_thread.is_alive():
                    break
                time.sleep(0.1)
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                self.ssh_config["hostname"],
                username=self.ssh_config["username"],
                password=self.ssh_config["password"]
            )
            # 只终止Auto_pack.sh相关进程
            kill_cmd = "ps aux | grep Auto_pack.sh | grep -v grep | awk '{print $2}' | xargs -r kill -9"
            ssh.exec_command(kill_cmd)
            ssh.close()
        except Exception as e:
            self.update_status(f"⚠️ 远程终止失败: {e}", "warning")
        self.restore_controls()
        self.status_badge.set_status("idle")
        # 延迟1.5秒后再显示终止成功提示，确保顺序正确且唯一
        self.app.after(1500, lambda: self.update_status("✅ 已终止切分任务，您可以继续操作。", "success"))

    def cleanup_current_dirs(self, ssh=None):
        import shutil
        # 本地清理
        for d in [getattr(self, '_current_nas_dir', None), getattr(self, '_current_out_dir', None)]:
            if d and os.path.exists(d):
                try:
                    shutil.rmtree(d)
                    self.update_status(f"🧹 已清理目录: {d}", "success")
                except Exception as e:
                    self.update_status(f"⚠️ 清理目录失败: {d} {e}", "warning")
        # 远程清理
        if ssh:
            for d in [getattr(self, '_current_nas_dir', None), getattr(self, '_current_out_dir', None)]:
                if d and d.startswith("/home/user/fkm/pack/"):
                    try:
                        ssh.exec_command(f"rm -rf {d}")
                        self.update_status(f"🧹 已远程清理目录: {d}", "success")
                    except Exception as e:
                        self.update_status(f"⚠️ 远程清理目录失败: {d} {e}", "warning")

    def _sftp_walk(self, sftp, remote_path):
        """递归遍历远程目录的辅助函数"""
        files = []
        folders = []
        
        # 确保使用正斜杠
        remote_path = remote_path.replace('\\', '/')
        
        for entry in sftp.listdir_attr(remote_path):
            if stat.S_ISDIR(entry.st_mode):
                folders.append(entry.filename)
            else:
                files.append(entry.filename)
            
        yield remote_path, folders, files
        
        for folder in folders:
            new_path = f"{remote_path}/{folder}"  # 使用正斜杠连接路径
            for x in self._sftp_walk(sftp, new_path):
                yield x

    def run(self):
        # 绑定窗口关闭事件
        self.app.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 设置窗口图标（如果有的话）
        try:
            # 这里可以设置窗口图标
            # self.app.iconbitmap("icon.ico")
            pass
        except:
            pass
            
        # 启动主循环
        try:
            self.app.mainloop()
        except Exception as e:
            print(f"程序运行出错: {e}")

    def on_closing(self):
        """窗口关闭处理"""
        if self.is_running:
            result = messagebox.askyesno(
                "确认", 
                "任务正在运行中，强制退出可能导致数据丢失。\n确定要退出吗？",
                icon='warning'
            )
            if not result:
                return
        
        try:
            # 停止所有后台任务
            if hasattr(self, 'app'):
                self.app.quit()
                self.app.destroy()
        except Exception as e:
            print(f"关闭程序时出错: {e}")
        finally:
            # 强制退出
            import sys
            sys.exit(0)

    def clear_status(self):
        """清空状态文本"""
        # 检查是否有内容
        content = self.status_text.get("1.0", "end-1c").strip()
        if not content:
            messagebox.showinfo("提示", "执行状态栏已为空！")
            return
            
        # 询问用户是否确定要清空
        if messagebox.askyesno("确认", "确定要清空所有日志吗？"):
            self.status_text.configure(state="normal")
            self.status_text.delete("1.0", "end")
            self.status_text.configure(state="disabled")
            self.status_badge.set_status("idle")
            messagebox.showinfo("成功", "日志已清空")

    def export_log(self):
        """导出日志到文件"""
        log_content = self.status_text.get("1.0", "end-1c").strip()
        if not log_content:
            messagebox.showwarning("警告", "没有日志可供导出！")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            title="导出日志"
        )
        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(log_content)
                messagebox.showinfo("成功", "日志导出成功！")
            except Exception as e:
                messagebox.showerror("错误", f"导出日志失败：{str(e)}")

    def on_terminate_button_click(self):
        """终止按钮点击逻辑：未运行时无响应，运行时弹窗确认"""
        if not self.is_running:
            return  # 未运行时无响应
        from tkinter import messagebox
        result = messagebox.askyesno("确认终止", "切分正在运行，确定终止当前任务吗？")
        if result:
            self.terminate_task()

# PyInstaller打包说明：
# pyinstaller -F -w -n pack --add-data "Auto_pack.sh;." --distpath ./pack-0715 切片工具数据备份/pack.py

if __name__ == "__main__":
    app = DataSlicerApp()
    app.run()